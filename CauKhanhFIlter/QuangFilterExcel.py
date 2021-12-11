from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import filedialog
import os

# Keyword input provided
a_condition = ['escalated to ASP', 'escalate to ASP', 'escalated ASP', 'Call ASP', 'Call depot', 'escalated depot', 'escalate to depot']
b_condition = ['Depot onsite', 'ASP onsite']

def start():
    print('Welcome to Quang\'s program!!')
    print('How to use:')
    print('1. Select your file to filter')
    print('2. Put data need filtering on K column')
    print('3. After finish select where to save your file')

    print('Aware: Here are the keywords use to filter data:')
    print('A. ' + str(a_condition))
    print('B. ' + str(b_condition))

    main_process()
    print('Finish!! Choose where to save your file!')
    print('Congrat!! Your file is successfully saved')
    print('Have fun! Press any key to quit!')


def filter_result( a : str  ) -> str :
    res = ""
    for i in a:
        if i.isnumeric() or i == ':':
            res += i
    return res 


def main_process():
    # ask for file path
    print('hello')
    root = tk.Tk()
    root.withdraw()
    file_input_path = filedialog.askopenfilename()
    print(file_input_path)
    print('Working on your file.... Please wait!')

    # Prepare variable
    wb = load_workbook(file_input_path)
    # add result Collumns:
    ws = wb['Sheet1']
    ws.insert_cols(12) #L
    ws.insert_cols(13) #M
    global a_condition
    global b_condition
    a_condition = [each_string.lower() for each_string in a_condition]
    b_condition = [each_string.lower() for each_string in b_condition]
    MAX_LENGTH = 10

    i = 1       #Line variable
    while ws['A'+ str(i)].value is not None:    # While first row still have val
        cellK = str(ws['K'+str(i)].value)
        line_in_cell = cellK.splitlines()
        result_a = ""
        result_b = ""
        for line in line_in_cell:
            line = line.lower()
            foundA = False

            for keyword_A in a_condition:
                # print(keyword_A)
                index_key = line.find(keyword_A)
                if index_key != -1:
                    add = filter_result(line[:MAX_LENGTH]) + ' \n'
                    result_a += add 
                    foundA = True
                    break
            if not foundA:
                for keyword_B in b_condition:
                    index_key = line.find(keyword_B)
                    if index_key != -1:
                        add = filter_result(line[:MAX_LENGTH]) + ' \n'
                        result_b += add 
                        
                        break
        ws['K' + str(i)].alignment = Alignment(wrap_text=True)
        ws['L' + str(i)].value = result_a
        ws['M' + str(i)].value = result_b
        i+=1        # Next line jump

    print('Please choose your file ! ')
    file_output_path = filedialog.askdirectory() + '/result.xlsx'
    print(file_output_path)
    wb.save(file_output_path)
    os.system("start EXCEL.EXE "+file_output_path)

try:
  start()
except Exception as e: print(e)

