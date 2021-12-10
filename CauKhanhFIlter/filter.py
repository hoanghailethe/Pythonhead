from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

wb = load_workbook('data.xlsx')
# add result Collumns:
ws = wb['Sheet1']
ws.insert_cols(12) #L
ws.insert_cols(13) #M

a_condition = ['escalated to ASP', 'escalate to ASP', 'escalated ASP', 'Call ASP', 'Call depot', 'escalated depot', 'escalate to depot']
b_condition = ['Depot onsite', 'ASP onsite']
a_condition = [each_string.lower() for each_string in a_condition]
b_condition = [each_string.lower() for each_string in b_condition]

def filter_result( a : str  ) -> str :
    res = ""
    for i in a:
        if i.isnumeric() or i == ':':
            res += i
    return res 

i = 1

while ws['A'+ str(i)].value is not None:
    cellK = str(ws['K'+str(i)].value)
    line_in_cell = cellK.splitlines()
    result_a = ""
    result_b = ""
    for line in line_in_cell:
        line = line.lower()
        foundA = False


        for keyword_A in a_condition:
            # print(keyword_A)
            index_key = line.find(keyword_A)
            if index_key != -1:
                add = filter_result(line[:index_key]) + '\n'
                result_a += add 
                foundA = True
                break
        if not foundA:
            for keyword_B in b_condition:
                index_key = line.find(keyword_B)
                if index_key != -1:
                    add = filter_result(line[:index_key]) + '\n'
                    result_b += add 
                    
                    break
    ws['K' + str(i)].alignment = Alignment(wrap_text=True)
    ws['L' + str(i)] = result_a
    ws['M' + str(i)] = result_b
    i+=1

wb.save('result.xlsx')