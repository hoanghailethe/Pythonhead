
import pandas as pd
from openpyxl import load_workbook

# assign directory
directory = 'D:/11.29.HH.moveLaptop/HHgithub/Pythonhead/Leetcode/IntegroFilterTTB/ExcelValFilter_branchID_CLOS_CLIMS'
 
# file input name
file_name = 'cifthieubranch.xlsx'


# Main process
df = pd.read_excel(open(directory + '/' +file_name, 'rb'),sheet_name='Sheet5')

# change type
# df['CIF 1'] = df['CIF 1'].astype(str)
# df['CIF 2'] = df['CIF 2'].astype(str)

#CIF 1 CASE CIF CÓ BÊN CLIMS VÀ CLOS
CIF1 = df['CIF 1']

#CIF2 CASE CIF CÓ BÊN CLIMS NHƯNG KHÔNG CÓ BÊN CLOS
CIF2 = df['CIF 2'].astype(float)

#CIF 3 CASE CIF CÓ BÊN CLOS, KHÔNG CÓ BÊN CLIMS
CIF3 = df['CIF 3'].astype(float)

print(CIF1)
print(CIF2)
print(CIF3)
print('--CIF--')

# inpu clos cms data
df_clos = pd.read_excel(open(directory + '/' +file_name, 'rb'),sheet_name='CLOS_in')
df_cms = pd.read_excel(open(directory + '/' +file_name, 'rb'),sheet_name='CMS_in')

df_cms['CMS_LE_SUB_PROFILE_ID'] = df_cms['CMS_LE_SUB_PROFILE_ID'].astype(str)
df_clos['ID'] = df_clos['ID'].astype(str)

#re1 - CMS CIF  res2 CLOS CIF
res1 = {}   #contain of clims
res2 = {}   #contain of clos

for cif in CIF1 : 
    for idx, x in enumerate(df_cms['LSP_LE_ID']) :
        if ( cif == x ):
            # res1[cif] = df_cms['CMS_SUB_ORIG_ORGANISATION'][idx]
            # continue
            res1[cif] = df_cms['CMS_LE_SUB_PROFILE_ID'][idx]
            continue

    for idy, y in enumerate(df_clos['CIF_NUMBER']) :
        if ( cif == y ):
            # res2[cif] = df_clos['BRANCH_ID'][idy]
            # continue
            res2[cif] = df_clos['ID'][idy]
            continue

for cif in CIF2 : 
    for idx, x in enumerate(df_cms['LSP_LE_ID']) :
        if ( cif == x ):
            # res3[cif] = df_cms['CMS_SUB_ORIG_ORGANISATION'][idx]
            # continue
            res1[cif] = df_cms['CMS_LE_SUB_PROFILE_ID'][idx]
            continue

for cif in CIF3 : 
    for idx, x in enumerate(df_clos['CIF_NUMBER']) :
        if ( cif == x ):
            # res3[cif] = df_cms['CMS_SUB_ORIG_ORGANISATION'][idx]
            # continue
            res2[cif] = df_clos['ID'][idx]
            continue

print(res1)
print('----')
print(res2)


myworkbook = load_workbook(directory + '/' +file_name)
worksheet= myworkbook.get_sheet_by_name('Sheet5')

for i in range(1, worksheet.max_row+1):
      
    # for j in range(1, worksheet.max_column+1):
    #     cell_obj = sh.cell(row=i, column=j)
    #     print(cell_obj.value, end=" ")

    cif_cms_and_clos_row_value = worksheet.cell(row = i, column=1).value
    cif_cms_row_value = worksheet.cell(row = i, column=5).value
    cif_clos_row_value = worksheet.cell(row = i, column=8).value


    # print("\n")
    # print("Row ", i, " data : " + str(cif_cms_row_value))

    if cif_cms_and_clos_row_value in res1:
        # print('ẼIST res1')
        worksheet.cell(row= i , column = 3 ).value = res1.get(cif_cms_and_clos_row_value)
        worksheet.cell(row= i , column = 2 ).value = res2.get(cif_cms_and_clos_row_value)
    if cif_cms_row_value in res1:
        worksheet.cell(row= i , column = 6 ).value = res1.get(cif_cms_row_value)

    if cif_clos_row_value in res2 :
        # print('ẼIST res3')
        worksheet.cell(row= i , column = 9 ).value = res2.get(cif_clos_row_value)

myworkbook.save(directory+'/'+'res.xlsx')
print('Write Successfully!')